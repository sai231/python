from openpyxl import Workbook
wb=Workbook()
ws=wb.active
ws['A1']=42
ws.append([1,2,3,4,5,6,7,8,9,10])
import datetime
ws['C1']=datetime.datetime.now()
ws1=wb.create_sheet(0)
ws.title="sai"
ws1.sheet_properties.tabColor="1072BA"
ws is wb
for sheet in wb:
    print(sheet.title)
for i in range(1,10):
    for j in range (1,101):
        wb.save("test.xlsx")
